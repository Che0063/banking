from fastapi import FastAPI, HTTPException, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
import sqlite3
import csv
import io
from datetime import datetime, timedelta
import re
import os
import secrets
from collections import Counter
from difflib import SequenceMatcher

app = FastAPI(title="Banking API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.add_middleware(GZipMiddleware)

DB_PATH = "/data/banking.db"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
_valid_tokens: set = set()
security = HTTPBearer(auto_error=False)

DEFAULT_CATEGORIES = [
    "Groceries","Food","Transport","Home","Health",
    "Clothing","Entertainment","Utilities","Income","Transfer","Other"
]

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            value_date TEXT,
            amount REAL NOT NULL,
            merchant TEXT NOT NULL,
            category TEXT NOT NULL,
            notes TEXT,
            person1_pct REAL,
            person2_pct REAL,
            is_transfer INTEGER DEFAULT 0,
            is_starting_balance INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            is_default INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pattern TEXT NOT NULL,
            category TEXT NOT NULL,
            use_regex INTEGER DEFAULT 0,
            enabled INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        INSERT OR IGNORE INTO settings VALUES ('person1_name','Person 1');
        INSERT OR IGNORE INTO settings VALUES ('person2_name','Person 2');
        INSERT OR IGNORE INTO settings VALUES ('person1_starting_balance','0');
        INSERT OR IGNORE INTO settings VALUES ('person2_starting_balance','0');
        CREATE TABLE IF NOT EXISTS import_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            imported_at TEXT DEFAULT (datetime('now')),
            source TEXT,
            count INTEGER
        );
        CREATE TABLE IF NOT EXISTS import_session_ids (
            session_id INTEGER,
            transaction_id INTEGER
        );
        INSERT OR IGNORE INTO settings VALUES ('dup_match_date','1');
        INSERT OR IGNORE INTO settings VALUES ('dup_match_amount','1');
        INSERT OR IGNORE INTO settings VALUES ('dup_match_merchant','1');
        INSERT OR IGNORE INTO settings VALUES ('dup_date_tolerance','3');
        INSERT OR IGNORE INTO settings VALUES ('dup_merchant_threshold','90');
    """)
    existing = {r[1] for r in conn.execute("PRAGMA table_info(transactions)").fetchall()}
    for col, typedef in [("value_date","TEXT"),("is_starting_balance","INTEGER DEFAULT 0")]:
        if col not in existing:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {typedef}")
    rule_cols = {r[1] for r in conn.execute("PRAGMA table_info(rules)").fetchall()}
    for col, typedef in [("enabled","INTEGER DEFAULT 1"),("sort_order","INTEGER DEFAULT 0")]:
        if col not in rule_cols:
            conn.execute(f"ALTER TABLE rules ADD COLUMN {col} {typedef}")
    # Migrate old single starting_balance to per-person
    old_rows = conn.execute("SELECT value FROM settings WHERE key='starting_balance'").fetchone()
    if old_rows:
        v = old_rows[0]
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('person1_starting_balance',?)", (v,))
        conn.execute("INSERT OR IGNORE INTO settings VALUES ('person2_starting_balance',?)", (v,))
        conn.execute("DELETE FROM settings WHERE key='starting_balance'")
    for cat in DEFAULT_CATEGORIES:
        conn.execute("INSERT OR IGNORE INTO categories (name,is_default) VALUES (?,1)", (cat,))
    conn.commit()
    conn.close()

init_db()

# ── Auth ──────────────────────────────────────────────────────────────────────
def check_auth(credentials: HTTPAuthorizationCredentials = Depends(security), token: str = None):
    if not APP_PASSWORD: return True
    # Accept token as query param for file downloads (browsers can't set headers for GET)
    tok = (credentials.credentials if credentials else None) or token
    if tok not in _valid_tokens:
        raise HTTPException(401, "Unauthorized")
    return True

@app.post("/api/auth/login")
def login(body: dict):
    if not APP_PASSWORD:
        t = "no-auth"; _valid_tokens.add(t); return {"token": t}
    if body.get("password") == APP_PASSWORD:
        t = secrets.token_hex(32); _valid_tokens.add(t); return {"token": t}
    raise HTTPException(401, "Incorrect password")

@app.get("/api/auth/check")
def auth_check(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not APP_PASSWORD: return {"ok": True, "password_required": False}
    if credentials and credentials.credentials in _valid_tokens:
        return {"ok": True, "password_required": True}
    return {"ok": False, "password_required": True}

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_date(raw: str) -> Optional[str]:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try: return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError: continue
    return None

def extract_value_date(desc: str) -> Optional[str]:
    m = re.search(r'Value Date:\s*(\d{2}/\d{2}/\d{4})', desc, re.IGNORECASE)
    return parse_date(m.group(1)) if m else None

def clean_merchant(desc: str) -> str:
    desc = re.sub(r'\s*Value Date:\s*\d{2}/\d{2}/\d{4}', '', desc, flags=re.IGNORECASE)
    desc = re.sub(r'\s+AUS Card xx\w+', '', desc)
    desc = re.sub(r'\s+AU$', '', desc)
    return desc.strip()

def get_categories_list(conn) -> List[str]:
    return [r["name"] for r in conn.execute(
        "SELECT name FROM categories ORDER BY is_default DESC, name ASC").fetchall()]

def apply_rules_to_merchant(merchant: str, conn) -> Optional[str]:
    rules = conn.execute(
        "SELECT * FROM rules WHERE enabled=1 ORDER BY sort_order ASC, id ASC").fetchall()
    for rule in rules:
        try:
            if rule["use_regex"]:
                if re.search(rule["pattern"], merchant, re.IGNORECASE): return rule["category"]
            else:
                if rule["pattern"].lower() in merchant.lower(): return rule["category"]
        except re.error: continue
    return None

def guess_category(desc: str, conn=None) -> str:
    if conn:
        matched = apply_rules_to_merchant(desc, conn)
        if matched: return matched
    d = desc.lower()
    if any(k in d for k in ["woolworths","coles","aldi","iga","market","lanka","saccas","supermarket"]): return "Groceries"
    if any(k in d for k in ["ptv","myki","uber","taxi","cab","fuel","bp ","petrol","atlas fuel"]): return "Transport"
    if any(k in d for k in ["fast transfer from","transfer from"]): return "Income"
    if any(k in d for k in ["transfer to","bpay","payto"]): return "Transfer"
    if any(k in d for k in ["medibank","bupa","pharmacy","chemist","clinic","hospital","dental","medical","physio","medicare"]): return "Health"
    if any(k in d for k in ["restaurant","cafe","mcdonald","kfc","hungry","pizza","sushi","subway","dominos","bakery"]): return "Food"
    if any(k in d for k in ["kmart","bunnings","ikea","officeworks","harvey"]): return "Home"
    if any(k in d for k in ["vodafone","exetel","optus","telstra"]): return "Utilities"
    if any(k in d for k in ["salary","wage","payroll","income"]): return "Income"
    return "Other"

def row_to_tx(row) -> dict:
    p1, p2, amt = row["person1_pct"], row["person2_pct"], row["amount"]
    pending = (p1 is None) and not row["is_transfer"] and not row["is_starting_balance"]
    return {
        "id": row["id"], "date": row["date"], "value_date": row["value_date"],
        "amount": amt, "merchant": row["merchant"], "category": row["category"],
        "notes": row["notes"], "person1_pct": p1, "person2_pct": p2,
        "person1_amount": round(p1 * amt, 2) if p1 is not None else None,
        "person2_amount": round(p2 * amt, 2) if p2 is not None else None,
        "is_transfer": bool(row["is_transfer"]),
        "is_starting_balance": bool(row["is_starting_balance"]),
        "is_pending": pending,
    }

def get_dup_config(conn) -> dict:
    rows = conn.execute("SELECT key,value FROM settings WHERE key LIKE 'dup_%'").fetchall()
    cfg = {r["key"]: r["value"] for r in rows}
    return {
        "match_date": cfg.get("dup_match_date","1") == "1",
        "match_amount": cfg.get("dup_match_amount","1") == "1",
        "match_merchant": cfg.get("dup_match_merchant","1") == "1",
        "date_tolerance": int(cfg.get("dup_date_tolerance","3")),
        "merchant_threshold": int(cfg.get("dup_merchant_threshold","90")),
    }

def merchant_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio() * 100

def find_duplicate(conn, date: str, amount: float, merchant: str, cfg: dict) -> Optional[dict]:
    """Return the first existing transaction that matches per config, or None."""
    if not (cfg["match_date"] or cfg["match_amount"] or cfg["match_merchant"]):
        return None
    rows = conn.execute("SELECT * FROM transactions WHERE is_starting_balance=0").fetchall()
    tol = cfg["date_tolerance"]
    try:
        incoming_dt = datetime.strptime(date, "%Y-%m-%d")
    except Exception:
        return None
    for r in rows:
        try:
            existing_dt = datetime.strptime(r["date"], "%Y-%m-%d")
        except Exception:
            continue
        date_ok = (not cfg["match_date"]) or abs((incoming_dt - existing_dt).days) <= tol
        amount_ok = (not cfg["match_amount"]) or abs(r["amount"] - amount) < 0.001
        merchant_ok = True
        if cfg["match_merchant"]:
            sim = merchant_similarity(r["merchant"], merchant)
            merchant_ok = sim >= cfg["merchant_threshold"]
        if date_ok and amount_ok and merchant_ok:
            return dict(r)
    return None

# ── Models ────────────────────────────────────────────────────────────────────
class TransactionIn(BaseModel):
    date: str; value_date: Optional[str] = None; amount: float
    merchant: str; category: str; notes: Optional[str] = None
    person1_pct: Optional[float] = None; is_transfer: Optional[bool] = False
    is_starting_balance: Optional[bool] = False

class BulkDeleteIn(BaseModel):
    ids: List[int]

class BulkEditIn(BaseModel):
    ids: List[int]; date: Optional[str] = None
    category: Optional[str] = None; person1_pct: Optional[float] = None

class RuleIn(BaseModel):
    pattern: str; category: str; use_regex: bool = False

# ── Transactions ──────────────────────────────────────────────────────────────
@app.get("/api/transactions", dependencies=[Depends(check_auth)])
def list_transactions(
    limit: int = 50, offset: int = 0,
    search: str = None, sort_col: str = "date", sort_dir: str = "desc",
    date_from: str = None, date_to: str = None,
    merchants: str = None, categories: str = None,
    select_all: bool = False
):
    conn = get_db()
    base = "FROM transactions WHERE 1=1"
    params = []
    if search:
        base += " AND (merchant LIKE ? OR notes LIKE ?)"; params += [f"%{search}%", f"%{search}%"]
    if date_from:
        base += " AND date>=?"; params.append(date_from)
    if date_to:
        base += " AND date<=?"; params.append(date_to)
    if merchants:
        mv = merchants.split("|")
        base += f" AND merchant IN ({','.join('?'*len(mv))})"; params += mv
    if categories:
        cv = categories.split("|")
        base += f" AND category IN ({','.join('?'*len(cv))})"; params += cv
    safe_cols = {"date","amount","merchant","category","notes","created_at"}
    sc = sort_col if sort_col in safe_cols else "date"
    sd = "ASC" if sort_dir.lower() == "asc" else "DESC"
    total = conn.execute(f"SELECT COUNT(*) {base}", params).fetchone()[0]
    if select_all:
        rows = conn.execute(f"SELECT id {base} ORDER BY {sc} {sd}", params).fetchall()
        conn.close()
        return {"transactions": [], "total": total, "all_ids": [r["id"] for r in rows]}
    rows = conn.execute(f"SELECT * {base} ORDER BY {sc} {sd} LIMIT ? OFFSET ?", params + [limit, offset]).fetchall()
    conn.close()
    return {"transactions": [row_to_tx(r) for r in rows], "total": total, "all_ids": []}

@app.get("/api/transactions/column-values", dependencies=[Depends(check_auth)])
def column_values(col: str):
    safe = {"merchant","category","date","value_date"}
    if col not in safe: raise HTTPException(400, "Invalid column")
    conn = get_db()
    rows = conn.execute(f"SELECT DISTINCT {col} FROM transactions ORDER BY {col} ASC").fetchall()
    conn.close()
    return [r[0] for r in rows if r[0]]

@app.post("/api/transactions", dependencies=[Depends(check_auth)])
def create_transaction(tx: TransactionIn):
    conn = get_db()
    p2 = round(1 - tx.person1_pct, 6) if tx.person1_pct is not None else None
    conn.execute(
        "INSERT INTO transactions (date,value_date,amount,merchant,category,notes,person1_pct,person2_pct,is_transfer,is_starting_balance) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (tx.date, tx.value_date, tx.amount, tx.merchant, tx.category, tx.notes,
         tx.person1_pct, p2, int(tx.is_transfer or False), int(tx.is_starting_balance or False))
    )
    conn.commit()
    row = conn.execute("SELECT * FROM transactions WHERE id=last_insert_rowid()").fetchone()
    conn.close()
    return row_to_tx(row)

@app.put("/api/transactions/{tx_id}", dependencies=[Depends(check_auth)])
def update_transaction(tx_id: int, tx: TransactionIn):
    conn = get_db()
    p2 = round(1 - tx.person1_pct, 6) if tx.person1_pct is not None else None
    conn.execute(
        "UPDATE transactions SET date=?,value_date=?,amount=?,merchant=?,category=?,notes=?,person1_pct=?,person2_pct=?,is_transfer=?,is_starting_balance=? WHERE id=?",
        (tx.date, tx.value_date, tx.amount, tx.merchant, tx.category, tx.notes,
         tx.person1_pct, p2, int(tx.is_transfer or False), int(tx.is_starting_balance or False), tx_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    conn.close()
    if not row: raise HTTPException(404, "Not found")
    return row_to_tx(row)

@app.delete("/api/transactions/bulk", dependencies=[Depends(check_auth)])
def bulk_delete(body: BulkDeleteIn):
    if not body.ids: return {"deleted": 0}
    conn = get_db()
    ph = ",".join("?" * len(body.ids))
    conn.execute(f"DELETE FROM transactions WHERE id IN ({ph})", body.ids)
    deleted = conn.total_changes
    conn.commit(); conn.close()
    return {"deleted": deleted}

@app.post("/api/transactions/bulk-delete", dependencies=[Depends(check_auth)])
def bulk_delete_post(body: BulkDeleteIn):
    """POST for bulk delete - DELETE+body is stripped by nginx."""
    if not body.ids: return {"deleted": 0}
    conn = get_db()
    ph = ",".join("?" * len(body.ids))
    conn.execute(f"DELETE FROM transactions WHERE id IN ({ph})", body.ids)
    deleted = conn.total_changes
    conn.commit(); conn.close()
    return {"deleted": deleted}

@app.put("/api/transactions/bulk-edit", dependencies=[Depends(check_auth)])
def bulk_edit(body: BulkEditIn):
    if not body.ids: return {"updated": 0}
    conn = get_db()
    sets, params = [], []
    if body.date is not None: sets.append("date=?"); params.append(body.date)
    if body.category is not None: sets.append("category=?"); params.append(body.category)
    if body.person1_pct is not None:
        sets.append("person1_pct=?"); params.append(body.person1_pct)
        sets.append("person2_pct=?"); params.append(round(1 - body.person1_pct, 6))
    if not sets: return {"updated": 0}
    ph = ",".join("?" * len(body.ids))
    conn.execute(f"UPDATE transactions SET {','.join(sets)} WHERE id IN ({ph})", params + body.ids)
    updated = conn.total_changes
    conn.commit(); conn.close()
    return {"updated": updated}

@app.delete("/api/transactions/{tx_id}", dependencies=[Depends(check_auth)])
def delete_transaction(tx_id: int):
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
    conn.commit(); conn.close()
    return {"ok": True}

# ── Summary ───────────────────────────────────────────────────────────────────

@app.get("/api/transactions/running-balance", dependencies=[Depends(check_auth)])
def running_balance():
    """Return all transactions with running shared/p1/p2 balances, chronological."""
    conn = get_db()
    s = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    p1_start = float(s.get("person1_starting_balance", 0))
    p2_start = float(s.get("person2_starting_balance", 0))
    rows = conn.execute(
        "SELECT id,date,amount,merchant,category,person1_pct,person2_pct,is_transfer,is_starting_balance "
        "FROM transactions WHERE is_starting_balance=0 ORDER BY date ASC, id ASC"
    ).fetchall()
    conn.close()
    shared_bal = p1_start + p2_start
    p1_bal = p1_start
    p2_bal = p2_start
    result = []
    for r in rows:
        amt = r["amount"]
        p1, p2 = r["person1_pct"], r["person2_pct"]
        shared_bal += amt
        if p1 is not None:
            p1_bal += round(p1 * amt, 2)
            p2_bal += round(p2 * amt, 2)
        result.append({
            "id": r["id"],
            "date": r["date"],
            "merchant": r["merchant"],
            "category": r["category"],
            "amount": amt,
            "shared_balance": round(shared_bal, 2),
            "person1_balance": round(p1_bal, 2),
            "person2_balance": round(p2_bal, 2),
        })
    return result

@app.get("/api/summary", dependencies=[Depends(check_auth)])
def get_summary(year: int = None, month: int = None, category: str = None):
    conn = get_db()
    s = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    p1_start = float(s.get("person1_starting_balance", 0))
    p2_start = float(s.get("person2_starting_balance", 0))
    q = "SELECT * FROM transactions WHERE is_starting_balance=0"
    params = []
    if year: q += " AND strftime('%Y',date)=?"; params.append(str(year))
    if month: q += " AND strftime('%m',date)=?"; params.append(f"{month:02d}")
    if category: q += " AND category=?"; params.append(category)
    q += " ORDER BY date ASC, id ASC"
    rows = conn.execute(q, params).fetchall()
    years = [r[0] for r in conn.execute(
        "SELECT DISTINCT strftime('%Y',date) y FROM transactions WHERE is_starting_balance=0 ORDER BY y DESC").fetchall()]
    conn.close()
    p1_bal = p2_bal = shared_bal = 0.0
    cat_totals = {}; cat_p1 = {}; cat_p2 = {}
    pending_count = 0; monthly = {}; monthly_person = {}
    for r in rows:
        amt = r["amount"]; shared_bal += amt
        p1, p2 = r["person1_pct"], r["person2_pct"]
        if p1 is None and not r["is_transfer"]: pending_count += 1
        elif p1 is not None:
            p1_bal += round(p1 * amt, 2); p2_bal += round(p2 * amt, 2)
        cat = r["category"]
        if cat not in ("Transfer","Income") and amt < 0:
            absamt = abs(amt)
            cat_totals[cat] = cat_totals.get(cat, 0) + absamt
            if p1 is not None:
                cat_p1[cat] = cat_p1.get(cat, 0) + round(p1 * absamt, 2)
                cat_p2[cat] = cat_p2.get(cat, 0) + round(p2 * absamt, 2)
        mk = r["date"][:7]
        if mk not in monthly: monthly[mk] = {"income": 0.0, "expenses": 0.0}
        if amt > 0: monthly[mk]["income"] += amt
        else: monthly[mk]["expenses"] += abs(amt)
        if cat == "Income" and p1 is not None and amt > 0:
            if mk not in monthly_person: monthly_person[mk] = {"p1": 0.0, "p2": 0.0}
            monthly_person[mk]["p1"] += round(p1 * amt, 2)
            monthly_person[mk]["p2"] += round(p2 * amt, 2)
    rnd = lambda d: {k: round(v, 2) for k, v in sorted(d.items(), key=lambda x: -x[1])}
    return {
        "shared_balance": round(shared_bal + p1_start + p2_start, 2),
        "person1_balance": round(p1_bal + p1_start, 2),
        "person2_balance": round(p2_bal + p2_start, 2),
        "pending_count": pending_count,
        "category_totals": rnd(cat_totals),
        "category_totals_p1": rnd(cat_p1),
        "category_totals_p2": rnd(cat_p2),
        "monthly": {k: {kk: round(vv,2) for kk,vv in v.items()} for k,v in sorted(monthly.items())},
        "monthly_person": {k: {kk: round(vv,2) for kk,vv in v.items()} for k,v in sorted(monthly_person.items())},
        "available_years": years,
    }

# ── Categories ────────────────────────────────────────────────────────────────
@app.get("/api/categories", dependencies=[Depends(check_auth)])
def get_categories():
    conn = get_db(); cats = get_categories_list(conn); conn.close(); return cats

@app.post("/api/categories", dependencies=[Depends(check_auth)])
def create_category(body: dict):
    name = body.get("name","").strip()
    if not name: raise HTTPException(400, "Name required")
    conn = get_db()
    try: conn.execute("INSERT INTO categories (name,is_default) VALUES (?,0)", (name,)); conn.commit()
    except sqlite3.IntegrityError: raise HTTPException(409, "Already exists")
    conn.close(); return {"ok": True}

@app.put("/api/categories/{old_name}", dependencies=[Depends(check_auth)])
def rename_category(old_name: str, body: dict):
    new_name = body.get("name","").strip()
    if not new_name: raise HTTPException(400, "Name required")
    conn = get_db()
    conn.execute("UPDATE categories SET name=? WHERE name=?", (new_name, old_name))
    conn.execute("UPDATE transactions SET category=? WHERE category=?", (new_name, old_name))
    conn.execute("UPDATE rules SET category=? WHERE category=?", (new_name, old_name))
    conn.commit(); conn.close(); return {"ok": True}

@app.delete("/api/categories/{name}", dependencies=[Depends(check_auth)])
def delete_category(name: str):
    conn = get_db()
    conn.execute("DELETE FROM categories WHERE name=?", (name,))
    conn.execute("UPDATE transactions SET category='Other' WHERE category=?", (name,))
    conn.commit(); conn.close(); return {"ok": True}

# ── Rules ─────────────────────────────────────────────────────────────────────
@app.get("/api/rules", dependencies=[Depends(check_auth)])
def list_rules():
    conn = get_db()
    rows = conn.execute("SELECT * FROM rules ORDER BY sort_order ASC, id ASC").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/rules", dependencies=[Depends(check_auth)])
def create_rule(rule: RuleIn):
    if rule.use_regex:
        try: re.compile(rule.pattern)
        except re.error as e: raise HTTPException(400, f"Invalid regex: {e}")
    conn = get_db()
    mo = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM rules").fetchone()[0]
    conn.execute("INSERT INTO rules (pattern,category,use_regex,sort_order) VALUES (?,?,?,?)",
                 (rule.pattern, rule.category, int(rule.use_regex), mo + 1))
    conn.commit()
    row = conn.execute("SELECT * FROM rules WHERE id=last_insert_rowid()").fetchone()
    conn.close(); return dict(row)

@app.put("/api/rules/{rule_id}", dependencies=[Depends(check_auth)])
def update_rule(rule_id: int, body: dict):
    conn = get_db()
    sets, params = [], []
    for f in ["pattern","category","use_regex","enabled","sort_order"]:
        if f in body: sets.append(f"{f}=?"); params.append(body[f])
    if sets: conn.execute(f"UPDATE rules SET {','.join(sets)} WHERE id=?", params + [rule_id]); conn.commit()
    conn.close(); return {"ok": True}

@app.delete("/api/rules/{rule_id}", dependencies=[Depends(check_auth)])
def delete_rule(rule_id: int):
    conn = get_db()
    conn.execute("DELETE FROM rules WHERE id=?", (rule_id,))
    conn.commit(); conn.close(); return {"ok": True}

@app.post("/api/rules/preview", dependencies=[Depends(check_auth)])
def preview_rule_apply(body: dict = {}):
    """Preview which transactions would change. scope: 'pending' | 'all'"""
    scope = body.get("scope", "pending")
    conn = get_db()
    if scope == "pending":
        rows = conn.execute(
            "SELECT id,merchant,category,date FROM transactions WHERE person1_pct IS NULL AND is_transfer=0 AND is_starting_balance=0"
        ).fetchall()
    else:
        rows = conn.execute("SELECT id,merchant,category,date FROM transactions").fetchall()
    changes = []
    for r in rows:
        new_cat = apply_rules_to_merchant(r["merchant"], conn)
        if new_cat and new_cat != r["category"]:
            changes.append({"id": r["id"], "merchant": r["merchant"],
                            "old_category": r["category"], "new_category": new_cat,
                            "date": r["date"]})
    conn.close(); return changes

@app.post("/api/rules/apply", dependencies=[Depends(check_auth)])
def apply_rules(body: dict):
    ids = body.get("ids", [])
    if not ids: return {"updated": 0}
    conn = get_db()
    updated = 0
    for tx_id in ids:
        row = conn.execute("SELECT merchant FROM transactions WHERE id=?", (tx_id,)).fetchone()
        if not row: continue
        cat = apply_rules_to_merchant(row["merchant"], conn)
        if cat: conn.execute("UPDATE transactions SET category=? WHERE id=?", (cat, tx_id)); updated += 1
    conn.commit(); conn.close(); return {"updated": updated}

@app.get("/api/rules/suggestions", dependencies=[Depends(check_auth)])
def rule_suggestions():
    conn = get_db()
    rows = conn.execute("SELECT merchant,category FROM transactions WHERE category!='Transfer'").fetchall()
    conn.close()
    mc = {}
    for r in rows:
        m = clean_merchant(r["merchant"])
        if len(m) < 3: continue
        if m not in mc: mc[m] = Counter()
        mc[m][r["category"]] += 1
    suggestions = []
    for merchant, counts in mc.items():
        top_cat, top_count = counts.most_common(1)[0]
        total = sum(counts.values())
        if total >= 2 and top_count / total >= 0.8:
            words = re.split(r'\s+', merchant.upper())
            keyword = " ".join(words[:2]) if len(words) >= 2 else words[0]
            if len(keyword) >= 4:
                suggestions.append({"pattern": keyword, "category": top_cat,
                                    "count": total, "confidence": round(top_count/total, 2)})
    seen = {}
    for s in sorted(suggestions, key=lambda x: -x["count"]):
        if s["pattern"] not in seen: seen[s["pattern"]] = s
    return sorted(seen.values(), key=lambda x: -x["count"])[:20]

# ── Settings ──────────────────────────────────────────────────────────────────
@app.get("/api/settings", dependencies=[Depends(check_auth)])
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    conn.close(); return {r["key"]: r["value"] for r in rows}

@app.put("/api/settings", dependencies=[Depends(check_auth)])
def update_settings(body: dict):
    conn = get_db()
    allowed = ["person1_name","person2_name","person1_starting_balance","person2_starting_balance",
               "dup_match_date","dup_match_amount","dup_match_merchant","dup_date_tolerance","dup_merchant_threshold"]
    for key in allowed:
        if key in body:
            conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", (key, str(body[key])))
    conn.commit(); conn.close(); return {"ok": True}

# ── Export ────────────────────────────────────────────────────────────────────
@app.get("/api/export/xlsx")
def export_xlsx(token: str = None, credentials: HTTPAuthorizationCredentials = Depends(security)):
    if APP_PASSWORD:
        tok = (credentials.credentials if credentials else None) or token
        if tok not in _valid_tokens:
            raise HTTPException(401, "Unauthorized")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    conn = get_db()
    s = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    p1_name = s.get("person1_name","Person 1")
    p2_name = s.get("person2_name","Person 2")
    p1_start = float(s.get("person1_starting_balance","0"))
    p2_start = float(s.get("person2_starting_balance","0"))

    rows = conn.execute(
        "SELECT * FROM transactions WHERE is_starting_balance=0 ORDER BY date ASC, id ASC"
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header row matching original format
    headers = [
        "Date", "Total Amount",
        f"% {p1_name}", f"% {p2_name}",
        f"$ {p1_name}", f"$ {p2_name}",
        "To/From", "Purpose/Notes", "Category",
        "Shared Balance", f"$ {p1_name} balance", f"$ {p2_name} balance"
    ]
    hdr_fill = PatternFill("solid", fgColor="2D6A4F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows with running balances
    shared_bal = p1_start + p2_start
    p1_bal = p1_start; p2_bal = p2_start
    alt_fill = PatternFill("solid", fgColor="F7F6F3")

    for ri, r in enumerate(rows, 2):
        amt = r["amount"]
        p1, p2 = r["person1_pct"], r["person2_pct"]
        p1_pct_disp = round(p1 * 100, 1) if p1 is not None else None
        p2_pct_disp = round(p2 * 100, 1) if p2 is not None else None
        p1_amt = round(p1 * amt, 2) if p1 is not None else None
        p2_amt = round(p2 * amt, 2) if p2 is not None else None
        shared_bal += amt
        if p1 is not None:
            p1_bal += p1_amt; p2_bal += p2_amt

        row_data = [
            r["date"], amt,
            p1_pct_disp, p2_pct_disp,
            p1_amt, p2_amt,
            r["merchant"], r["notes"], r["category"],
            round(shared_bal,2), round(p1_bal,2), round(p2_bal,2)
        ]
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if fill: cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            # Format numbers
            if ci == 2: cell.number_format = '#,##0.00'
            if ci in (3,4): cell.number_format = '0.0'
            if ci in (5,6,10,11,12): cell.number_format = '#,##0.00'
            if ci in (5,6,2) and val is not None:
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    col_widths = [12,14,10,10,12,12,30,30,14,14,16,16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    filename = f"banking_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ── Import ────────────────────────────────────────────────────────────────────
def parse_rows_commbank(content_bytes: bytes, conn) -> tuple:
    text = content_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    pending = []; skipped = 0
    cfg = get_dup_config(conn)
    for row in reader:
        if len(row) < 3 or row[0].strip().lower() in ("date",""): continue
        try:
            desktop_date = parse_date(row[0].strip())
            if not desktop_date: skipped += 1; continue
            amount = float(row[1].strip().replace('"','').replace('+',''))
            description = row[2].strip()
            value_date = extract_value_date(description)
            merchant = clean_merchant(description) or description
            category = guess_category(description, conn)
            is_transfer = 1 if category == "Transfer" else 0
            dup = find_duplicate(conn, desktop_date, amount, merchant, cfg)
            pending.append({
                "date": desktop_date, "value_date": value_date, "amount": amount,
                "merchant": merchant, "category": category, "is_transfer": is_transfer,
                "duplicate_of": dup,
            })
        except Exception: skipped += 1
    return pending, skipped

def parse_rows_xlsx(content_bytes: bytes, conn) -> tuple:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    ws = wb.active
    pending = []; skipped = 0
    cfg = get_dup_config(conn)
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            raw_date = row[0]
            if raw_date is None: continue
            if isinstance(raw_date, datetime): iso_date = raw_date.strftime("%Y-%m-%d")
            elif isinstance(raw_date, str):
                iso_date = parse_date(raw_date)
                if not iso_date: skipped += 1; continue
            else: skipped += 1; continue
            raw_amount = row[1]
            if raw_amount is None or isinstance(raw_amount, str): skipped += 1; continue
            amount = float(raw_amount)
            p1_raw = row[2]
            p1_pct = float(p1_raw) if isinstance(p1_raw, (int, float)) else None
            p2_pct = round(1 - p1_pct, 6) if p1_pct is not None else None
            merchant = str(row[6]).strip() if row[6] else "Unknown"
            if merchant in ("None","Unknown",""): skipped += 1; continue
            notes_val = row[7]
            notes = str(notes_val).strip() if notes_val else None
            if notes in ("START","None",""): notes = None
            cat_val = row[8]
            category = str(cat_val).strip() if cat_val else "Other"
            if not category or category == "None": category = "Other"
            is_transfer = 1 if category == "Transfer" else 0
            dup = find_duplicate(conn, iso_date, amount, merchant, cfg)
            pending.append({
                "date": iso_date, "value_date": None, "amount": amount,
                "merchant": merchant, "category": category, "notes": notes,
                "person1_pct": p1_pct, "person2_pct": p2_pct,
                "is_transfer": is_transfer, "duplicate_of": dup,
            })
        except Exception: skipped += 1
    return pending, skipped

@app.post("/api/import/commbank/parse", dependencies=[Depends(check_auth)])
async def parse_commbank(file: UploadFile = File(...)):
    content = await file.read()
    conn = get_db()
    pending, skipped = parse_rows_commbank(content, conn)
    # Ensure categories exist
    for p in pending:
        conn.execute("INSERT OR IGNORE INTO categories (name,is_default) VALUES (?,0)", (p["category"],))
    conn.commit(); conn.close()
    return {"rows": pending, "skipped": skipped}

@app.post("/api/import/xlsx/parse", dependencies=[Depends(check_auth)])
async def parse_xlsx_import(file: UploadFile = File(...)):
    try: import openpyxl
    except ImportError: raise HTTPException(500, "openpyxl not installed")
    content = await file.read()
    conn = get_db()
    pending, skipped = parse_rows_xlsx(content, conn)
    for p in pending:
        conn.execute("INSERT OR IGNORE INTO categories (name,is_default) VALUES (?,0)", (p["category"],))
    conn.commit(); conn.close()
    return {"rows": pending, "skipped": skipped}

class ImportConfirmIn(BaseModel):
    rows: List[dict]  # each row has all fields + action: 'import'|'skip'

@app.post("/api/import/confirm", dependencies=[Depends(check_auth)])
def confirm_import(body: ImportConfirmIn):
    conn = get_db()
    imported = skipped = 0
    tx_ids = []
    source = body.rows[0].get("_source","unknown") if body.rows else "unknown"
    for row in body.rows:
        if row.get("action") == "skip": skipped += 1; continue
        try:
            conn.execute(
                "INSERT INTO transactions (date,value_date,amount,merchant,category,notes,person1_pct,person2_pct,is_transfer) VALUES (?,?,?,?,?,?,?,?,?)",
                (row.get("date"), row.get("value_date"), row.get("amount"),
                 row.get("merchant"), row.get("category"), row.get("notes"),
                 row.get("person1_pct"), row.get("person2_pct"),
                 int(row.get("is_transfer", 0)))
            )
            tx_ids.append(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            imported += 1
        except Exception: skipped += 1
    if tx_ids:
        conn.execute("INSERT INTO import_sessions (source, count) VALUES (?,?)", (source, len(tx_ids)))
        session_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for tid in tx_ids:
            conn.execute("INSERT INTO import_session_ids VALUES (?,?)", (session_id, tid))
    conn.commit(); conn.close()
    return {"imported": imported, "skipped": skipped}

@app.get("/api/import/sessions", dependencies=[Depends(check_auth)])
def list_import_sessions():
    conn = get_db()
    rows = conn.execute("SELECT * FROM import_sessions ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.delete("/api/import/sessions/{session_id}", dependencies=[Depends(check_auth)])
def undo_import_session(session_id: int):
    conn = get_db()
    ids = [r[0] for r in conn.execute(
        "SELECT transaction_id FROM import_session_ids WHERE session_id=?", (session_id,)).fetchall()]
    if ids:
        ph = ",".join("?" * len(ids))
        conn.execute(f"DELETE FROM transactions WHERE id IN ({ph})", ids)
    conn.execute("DELETE FROM import_session_ids WHERE session_id=?", (session_id,))
    conn.execute("DELETE FROM import_sessions WHERE id=?", (session_id,))
    conn.commit(); conn.close()
    return {"deleted": len(ids)}
